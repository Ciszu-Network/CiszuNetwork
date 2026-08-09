# -*- coding: utf-8 -*-
"""Genera la estructura legal de archives/legal (carpetas + tablas CSV + Excel de contabilidad).

Uso:
    python tools/legal-ai/generate-legal-structure.py

Qué crea (todo en archives/legal/, gitignored por privacidad):
    contabilidad/          — libro-ingresos.csv, libro-gastos.csv, control-impuestos.csv,
                             vencimientos.csv y contabilidad-ciszu-<AÑO>.xlsx (con fórmulas)
    evidencia-marca/       — registro-evidencia.csv (log de capturas de uso de la marca)
    documentos-oficiales/  — carpeta para cédula, RIF, actas, certificados (vacía)
    tramites/              — seguimiento-tramites.md (estado del plan por fases)

Los CSV son la fuente editable (Excel los abre con UTF-8 BOM); el XLSX es derivado:
    python tools/legal-ai/generate-legal-structure.py
"""
import os, csv, sys

ROOT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "archives", "legal"))
CONT = os.path.join(ROOT, "contabilidad")
EVID = os.path.join(ROOT, "evidencia-marca")
DOCS = os.path.join(ROOT, "documentos-oficiales")
TRAM = os.path.join(ROOT, "tramites")
for d in (CONT, EVID, DOCS, TRAM):
    os.makedirs(d, exist_ok=True)

# Año por defecto: actual; se puede pasar el año como argumento (--year 2026)
YEAR = str(__import__("datetime").date.today().year)
if "--year" in sys.argv:
    YEAR = sys.argv[sys.argv.index("--year") + 1]

def write_csv(path, headers, rows, note=None):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        if note:
            w.writerow(["# " + note])
        w.writerow(headers)
        w.writerows(rows)
    print("CSV OK", os.path.relpath(path, ROOT))

# ── 1. Libro de ingresos (base ISLR/IVA/IGTF y exportación de servicios) ──
write_csv(os.path.join(CONT, "libro-ingresos.csv"),
    ["Fecha", "Cliente/Plataforma", "Concepto", "País (VE/Extranjero)", "Moneda",
     "Monto USD", "Tasa BCV (Bs/USD)", "Monto Bs", "IVA (0% export/16%)", "IGTF 3%",
     "Total Bs", "Factura N°", "Método de pago", "RIF/ID cliente", "Notas"],
    [["2026-08-10", "[CLIENTE EJEMPLO — borrar]", "[Concepto ejemplo]", "Extranjero", "USD",
      "100.00", "38.50", "3850.00", "0% (exportación)", "115.50", "3965.50", "[N°]",
      "Payoneer", "[RIF/ID]", "Fila de ejemplo: reemplazar o borrar"]],
    note="Registrar cada ingreso el mismo día. Exportación de servicios digitales = 0% IVA. IGTF aplica a pagos recibidos del exterior.")

# ── 2. Libro de gastos (deducibles ISLR) ──
write_csv(os.path.join(CONT, "libro-gastos.csv"),
    ["Fecha", "Proveedor", "Concepto", "Categoría", "Comprobante N°", "Moneda",
     "Monto USD", "Tasa BCV (Bs/USD)", "Monto Bs", "¿Deducible ISLR? (Sí/No)", "Notas"],
    [["2026-08-10", "[PROVEEDOR EJEMPLO — borrar]", "[Concepto ejemplo]", "Hosting", "[N°]",
      "USD", "10.00", "38.50", "385.00", "Sí", "Fila de ejemplo: reemplazar o borrar"]],
    note="Guardar el comprobante (factura/recibo) en documentos-oficiales/ o con el nombre en 'Comprobante N°'.")

# ── 3. Control de impuestos (calendario y estimaciones) ──
write_csv(os.path.join(CONT, "control-impuestos.csv"),
    ["Periodo (MM-AAAA)", "Ingresos Bs", "IVA causado 16% (ventas locales)", "IVA exportación (0%)",
     "Retenciones IVA", "Retenciones ISLR", "IGTF pagado", "ISLR estimado",
     "Fecha decl. IVA", "Fecha decl. ISLR", "Estado"],
    [["2026-08", "", "", "", "", "", "", "", "[15 sep]", "[ene-mar 2027]", "Pendiente"]],
    note="Llenar mensualmente. ISLR: declaración anual 1 ene–31 mar del año siguiente.")

# ── 4. Vencimientos (no perder plazos = no pagar multas) ──
write_csv(os.path.join(CONT, "vencimientos.csv"),
    ["Trámite", "Órgano", "Frecuencia", "Próximo vencimiento", "Dónde se hace", "Estado"],
    [
        ["Mayoría de edad (18 años) — gatillo de trámites", "—", "Única", "[FECHA CUMPLEAÑOS]", "—", "Preparación: tener documentación lista"],
        ["RIF persona natural — inscripción inicial", "SENIAT", "Única (renovar cada 3 años)", "[al cumplir 18 años]", "declaraciones.seniat.gob.ve", "Preparación: documentación lista"],
        ["Declaración IVA (si aplica)", "SENIAT", "Mensual", "[día según calendario]", "declaraciones.seniat.gob.ve", "Pendiente"],
        ["Declaración ISLR anual", "SENIAT", "Anual (1 ene–31 mar)", "[marzo 2027]", "declaraciones.seniat.gob.ve", "Pendiente"],
        ["Facturación digital (Prov. 102/121)", "SENIAT", "Antes de facturar", "[inmediato si hay ingresos]", "declaraciones.seniat.gob.ve", "Pendiente"],
        ["Marca SAPI — búsqueda antecedentes", "SAPI", "Fase 1 (~$100)", "[primeras ganancias / al cumplir 18]", "sapi.gob.ve (WEBPI)", "Pendiente"],
        ["Marca SAPI — derechos registro", "SAPI", "30 días hábiles tras concesión", "[cuando se conceda]", "sapi.gob.ve (WEBPI)", "Pendiente"],
    ],
    note="Copiar estas fechas a un calendario con avisos.")

# ── 5. Registro de evidencias de uso de la marca (Fase 0 → Fase 1) ──
write_csv(os.path.join(EVID, "registro-evidencia.csv"),
    ["Fecha", "Plataforma/URL", "Qué muestra", "Tipo (web/bot/red/otro)", "Archivo adjunto", "Notas"],
    [["2026-08-10", "ciszunetwork.vercel.app", "Captura de la web principal", "web", "[ruta o nombre del archivo]", "Ejemplo de fila"]],
    note="Captura con fecha visible (sistema o URL). Nombrar: AAAAMMDD-plataforma-descripcion.png")

print("\nDirectorio:", ROOT)

# ── 6. Excel con fórmulas ──
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    def add_sheet(name, headers, rows, widths, formulas=None):
        ws = wb.active if name == "Ingresos" else wb.create_sheet(name)
        ws.title = name
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for r, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(row=r, column=c, value=v)
        if formulas:
            for r, expr in formulas:
                ws.cell(row=r, column=1, value=expr)
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    add_sheet("Ingresos",
        ["Fecha", "Cliente/Plataforma", "Concepto", "País", "Moneda", "Monto USD",
         "Tasa BCV", "Monto Bs", "IVA", "IGTF", "Total Bs", "Factura N°", "Pago", "RIF/ID", "Notas"],
        [["2026-08-10", "[CLIENTE EJEMPLO — borrar]", "[Concepto]", "Extranjero", "USD",
          100, 38.5, 3850, "0%", 115.5, 3965.5, "[N°]", "Payoneer", "[ID]", "Ejemplo: borrar"]],
        [12, 24, 28, 14, 8, 11, 11, 11, 14, 9, 11, 11, 12, 12, 30],
        [(1001, "=SUM(K2:K1000)")])

    add_sheet("Gastos",
        ["Fecha", "Proveedor", "Concepto", "Categoría", "Comprobante N°", "Moneda",
         "Monto USD", "Tasa BCV", "Monto Bs", "¿Deducible ISLR?", "Notas"],
        [["2026-08-10", "[PROVEEDOR EJEMPLO — borrar]", "[Concepto]", "Hosting", "[N°]",
          "USD", 10, 38.5, 385, "Sí", "Ejemplo: borrar"]],
        [12, 24, 28, 14, 15, 8, 11, 11, 11, 15, 30],
        [(1001, "=SUM(I2:I1000)")])

    add_sheet("Impuestos",
        ["Periodo", "Ingresos Bs", "IVA 16%", "IVA 0% exp.", "Ret. IVA", "Ret. ISLR",
         "IGTF", "ISLR estimado", "Fecha decl. IVA", "Fecha decl. ISLR", "Estado"],
        [["2026-08", None, None, None, None, None, None, None, "", "", "Pendiente"]],
        [12, 12, 11, 12, 10, 10, 10, 13, 15, 15, 12])

    add_sheet("Vencimientos",
        ["Trámite", "Órgano", "Frecuencia", "Próximo vencimiento", "Dónde se hace", "Estado"],
        [
            ["Mayoría de edad (18 años) — gatillo de trámites", "—", "Única", "", "—", "Preparación: tener documentación lista"],
            ["RIF persona natural — inscripción inicial", "SENIAT", "Única (renovar cada 3 años)", "", "declaraciones.seniat.gob.ve", "Preparación: documentación lista"],
            ["Declaración IVA (si aplica)", "SENIAT", "Mensual", "", "declaraciones.seniat.gob.ve", "Pendiente"],
            ["Declaración ISLR anual", "SENIAT", "Anual (1 ene–31 mar)", "", "declaraciones.seniat.gob.ve", "Pendiente"],
            ["Facturación digital (Prov. 102/121)", "SENIAT", "Antes de facturar", "", "declaraciones.seniat.gob.ve", "Pendiente"],
            ["Marca SAPI — búsqueda antecedentes", "SAPI", "Fase 1 (~$100)", "", "sapi.gob.ve (WEBPI)", "Pendiente"],
            ["Marca SAPI — derechos registro", "SAPI", "30 días hábiles tras concesión", "", "sapi.gob.ve (WEBPI)", "Pendiente"],
        ],
        [42, 10, 26, 20, 34, 26])

    add_sheet("Evidencia",
        ["Fecha", "Plataforma/URL", "Qué muestra", "Tipo", "Archivo adjunto", "Notas"],
        [["2026-08-10", "ciszunetwork.vercel.app", "Captura de la web principal", "web", "", "Ejemplo"]],
        [12, 30, 34, 14, 26, 30])

    xlsx = os.path.join(CONT, f"contabilidad-ciszu-{YEAR}.xlsx")
    wb.save(xlsx)
    print("XLSX OK", os.path.relpath(xlsx, ROOT))
except ImportError:
    print("XLSX SKIP: instalar openpyxl con: pip install openpyxl")
except Exception as e:
    print("XLSX FAIL:", e)
