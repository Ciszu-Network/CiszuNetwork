# -*- coding: utf-8 -*-
"""xlsx2csv — XLSX a CSV con openpyxl (una hoja por CSV: <base>-<hoja>.csv).
Uso: python xlsx2csv.py <input.xlsx> --out-dir <directorio>"""
import sys, os
from openpyxl import load_workbook

def main():
    args = sys.argv[1:]
    inp = args[0] if args else None
    out_dir = None
    if "--out-dir" in args:
        out_dir = args[args.index("--out-dir") + 1]
    if not inp or not out_dir:
        print("Uso: python xlsx2csv.py <input.xlsx> --out-dir <directorio>"); sys.exit(1)
    os.makedirs(out_dir, exist_ok=True)
    wb = load_workbook(inp, read_only=True)
    base = os.path.splitext(os.path.basename(inp))[0]
    for ws in wb.worksheets:
        out = os.path.join(out_dir, f"{base}-{ws.title}.csv")
        with open(out, "w", encoding="utf-8-sig", newline="") as f:
            import csv
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                w.writerow(["" if v is None else v for v in row])
        print(f"OK {out}")
    print(f"Hojas: {wb.sheetnames}")

if __name__ == "__main__":
    main()
