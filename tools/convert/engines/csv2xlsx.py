# -*- coding: utf-8 -*-
"""csv2xlsx — CSV a XLSX con openpyxl (headers en negrita, anchos automáticos).
Uso: python csv2xlsx.py <input.csv> --out <output.xlsx>"""
import csv, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def main():
    args = sys.argv[1:]
    inp = args[0] if args else None
    out = None
    if "--out" in args:
        out = args[args.index("--out") + 1]
    if not inp or not out:
        print("Uso: python csv2xlsx.py <input.csv> --out <output.xlsx>"); sys.exit(1)
    with open(inp, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    fill = PatternFill("solid", fgColor="1F2937")
    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if r == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
    for c in range(1, len(rows[0]) + 1 if rows else 1):
        ws.column_dimensions[get_column_letter(c)].width = max(len(str(rows[r][c - 1])) + 2 if len(rows[r]) >= c else 8 for r in range(len(rows))) + 2
    wb.save(out)
    print(f"OK {out} ({len(rows)} filas)")

if __name__ == "__main__":
    main()
